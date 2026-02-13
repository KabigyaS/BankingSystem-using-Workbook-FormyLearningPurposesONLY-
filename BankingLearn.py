from openpyxl import load_workbook, Workbook
import os
#AI was used for learning purposes
filename = input("Enter a name for your bank file (example: Accounts.xlsx): ")

folder = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(folder, filename)

if os.path.exists(path):
    wb = load_workbook(path)
else:
    wb = Workbook()
    wb.active.append(["Name", "Password", "Balance", "Transactions"])

sheet = wb.active


def checkBalance(name):
    for row in sheet.iter_rows(min_row=1,values_only=False):
        if row[0].value == name:
            print("Your balance: ", row[2].value)

def depositMoney(name,amount):
    for row in sheet.iter_rows(min_row=2,values_only=False):
        if row[0].value == name:
            row[2].value+=amount
            if row[3].value is None:
                row[3].value = ""
            row[3].value += f"\n Deposited {amount}"

def withdrawMoney(name,amount):
    for row in sheet.iter_rows(min_row=2,values_only=False):
        if row[0].value == name:
            row[2].value-=amount
            if row[3].value is None:
                row[3].value = ""
            row[3].value += f"\n Withdrawed {amount}"


def viewTransaction(name):
    for row in sheet.iter_rows(min_row=2,values_only=False):
        if row[0].value == name:
            if row[3].value == "" or row[3].value is None:
                print("No Transaction History")
            else:    
                print(row[3].value)

    
def newMember(name,password,amount):
    sheet.append([name,password,amount,f"first Added {amount}"])

def passwordCheck(name,password):
    for row in sheet.iter_rows(min_row=2,values_only=False):
        if row[0].value == name:
            if password == row[1].value:
                return True
            else:
                return False
        
def userFound(name):
    for row in sheet.iter_rows(min_row=2,values_only=False):
        if row[0].value == name:
            return True
    return False

print("Welcome to the Banking System \nwhat would you like to do today:")
print("1.Check Balance \n2.Deposit Money \n3.Withdraw money \n4.Transfer Money \n5.view transaction history \n6.Become a new member \n7.Exit")
option = int(input("Which option do you choose, enter the number: "))
name= input("What is your name: ")

if option != 6:
    while not userFound(name):
        print("User not Found")
        name= input("What is your name: ")
            
if option != 6:
    password=input("Enter your password: ")
    while not passwordCheck(name,password):
        print("Wrong Password")
        password=input("Enter your password: ")
else:
    password=input("Create a password:")

amount=0

while option < 7:
    try:
        if option == 1:
            checkBalance(name)

        elif option == 2:
            amount=int(input("Enter the amount you want deposit: "))
            depositMoney(name,amount)
            amount=0

        elif option == 3:
            amount=int(input("Enter the amount you want Withdraw: "))
            withdrawMoney(name,amount)
            amount=0

        elif option == 4:
            name2=input("Who do you want to transfer it to?(Name): ")
            amount=int(input("Enter the amount you want Transfer: "))
            depositMoney(name2,amount)
            withdrawMoney(name,amount)
            amount=0

        elif option == 5:
            viewTransaction(name)

        elif option == 6:
            x = [row[0] for row in sheet.iter_rows(min_row=2,values_only=True)]

            if name in x:
                print("Username already exist")
            else:
                amount=int(input("Enter the amount you want put into this bank: "))
                newMember(name,password,amount)

        option = int(input("Anything else, Which option do you choose, enter the number: "))

    except ValueError:
        print("invalid input....")

wb.save(path)

print(f"\nYour bank file is saved here:\n{path}")

